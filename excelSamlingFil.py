import datetime 
import pandas as pd
from openpyxl import Workbook, load_workbook


def call_main(event=None, context=None):
    date = datetime.datetime.now()

    filename="akskjestigninger-" + date.strftime('%Y-%m-%d') + ".xlsx"

    stocks = load_workbook(filename)
    stocks_sheet = stocks.active 

    all_stocks = load_workbook(filename="Investtech_Aksjestigning.xlsx")
    all_stocks_sheet = all_stocks.active 

    for row in stocks_sheet.iter_rows(min_row=2, values_only=True):
        all_stocks_sheet.append([
            date.strftime("%A, %B %d, %Y"), row[0], row[1], row[2], row[3], row[4], row[5]
        ])

    all_stocks.save('Investtech_Aksjestigning.xlsx')
